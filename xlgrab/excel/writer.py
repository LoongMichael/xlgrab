"""
Excel 写入模块

提供向现有Excel文件高效写入数据的功能。对外暴露三个核心函数：
- to_sheet_many: (推荐) 自动分批写入多个任务，性能最高。
- write_to_excel: 写入单个DataFrame。
- write_range_to_excel: 写入二维列表或元组。
"""

from itertools import groupby
from operator import itemgetter
from typing import Any, Dict, List, Optional, Union
import warnings

import openpyxl
import pandas as pd

# ====================================================================
# Public API
# ====================================================================

def to_sheet_many(tasks: List[Dict[str, Any]]) -> None:
    """
    自动按文件名分批，向多个Excel文件高效写入数据。

    本函数会自动将任务按 `excel_name` 分组，并对每个文件使用一次性的
    “打开-写入-保存”会话，是执行批量写入的最高效、最简洁的方式。

    参数:
        tasks: 字典列表。每个字典必须包含 'excel_name' 和 'df'，
               以及 `write_to_excel` 的其他可选参数。
               示例:
                   [
                       {"excel_name": "file1.xlsx", "df": df1, "sheet_name": "Data1"},
                       {"excel_name": "file2.xlsx", "df": df2, "sheet_name": "Data2"},
                       {"excel_name": "file1.xlsx", "df": df3, "sheet_name": "Data3"},
                   ]
    """
    # 为了让 groupby 生效，必须先按分组键排序
    sorted_tasks = sorted(tasks, key=itemgetter('excel_name'))

    # 按 excel_name 分组并处理每个文件
    for excel_name, group in groupby(sorted_tasks, key=itemgetter('excel_name')):
        with _ExcelBatchWriter(str(excel_name)) as writer:
            task_list = list(group)
            # 准备每个文件的写入任务（移除 excel_name 键）
            write_tasks = [{k: v for k, v in task.items() if k != 'excel_name'} for task in task_list]
            writer.write_many(write_tasks)


def write_to_excel(df: pd.DataFrame,
                   excel_name: str,
                   sheet_name: Union[str, int] = 0,
                   start_row: int = 1,
                   start_col: int = 1,
                   end_row: Optional[int] = None,
                   end_col: Optional[int] = None,
                   overwrite: bool = False,
                   header: bool = True,
                   index: bool = False,
                   *,
                   _workbook: Optional[openpyxl.Workbook] = None,
                   _save: bool = True) -> None:
    """
    向现有Excel文件的指定位置写入DataFrame数据。
    当单独使用时，每次调用都会打开和保存文件。推荐使用 `to_sheet_many` 进行多次写入。
    """

    if not isinstance(df, pd.DataFrame):
        raise ValueError("df参数必须是pandas DataFrame")
    if not isinstance(excel_name, str):
        raise ValueError("excel_name参数必须是字符串")
    if start_row < 1 or start_col < 1:
        raise ValueError("start_row/start_col必须大于等于1")

    # --- 核心写入逻辑 ---
    _perform_write(
        df, excel_name, sheet_name, start_row, start_col,
        end_row, end_col, overwrite, header, index, _workbook, _save
    )


def write_range_to_excel(data: Union[pd.DataFrame, list, tuple],
                         excel_name: str,
                         sheet_name: Union[str, int] = 0,
                         start_row: int = 1,
                         start_col: int = 1,
                         end_row: Optional[int] = None,
                         end_col: Optional[int] = None) -> None:
    """向Excel文件的指定范围写入数据（简化版本）。"""
    if isinstance(data, pd.DataFrame):
        df = data
    elif isinstance(data, (list, tuple)):
        df = pd.DataFrame(data)
    else:
        raise ValueError("data参数必须是DataFrame、列表或元组")

    write_to_excel(
        df=df,
        excel_name=excel_name,
        sheet_name=sheet_name,
        start_row=start_row,
        start_col=start_col,
        end_row=end_row,
        end_col=end_col,
        overwrite=True,
        header=False,
        index=False,
    )


# ====================================================================
# Internal Implementation
# ====================================================================

class _ExcelBatchWriter:
    """内部类：一次打开、多次写、一次保存，避免重复I/O。"""

    def __init__(self, excel_name: str):
        self.excel_name = excel_name
        self.workbook = _open_or_create_workbook(excel_name)

    def write(self, **kwargs) -> None:
        write_to_excel(
            excel_name=self.excel_name,
            _workbook=self.workbook,
            _save=False,
            **kwargs
        )

    def write_many(self, tasks: List[Dict[str, Any]]) -> None:
        for task in tasks:
            self.write(**task)

    def save(self) -> None:
        self.workbook.save(self.excel_name)

    def __enter__(self):
        return self

    def __exit__(self, exc_type, exc, tb):
        self.save()


def _perform_write(df: pd.DataFrame,
                   excel_name: str,
                   sheet_name: Union[str, int],
                   start_row: int,
                   start_col: int,
                   end_row: Optional[int],
                   end_col: Optional[int],
                   overwrite: bool,
                   header: bool,
                   index: bool,
                   _workbook: Optional[openpyxl.Workbook],
                   _save: bool) -> None:
    """包含所有写入逻辑的内部函数。"""
    # 实际尺寸
    df_rows, df_cols = len(df), len(df.columns)

    # 计算目标区域尺寸（如未指定，则按df尺寸）
    if end_row is None:
        end_row = start_row + df_rows - 1
    if end_col is None:
        end_col = start_col + df_cols - 1
    if end_row < start_row or end_col < start_col:
        raise ValueError("end_row/end_col不能小于start_row/start_col")

    required_rows = end_row - start_row + 1
    required_cols = end_col - start_col + 1

    # 截断/填充提示
    if not overwrite and (df_rows > required_rows or df_cols > required_cols):
        warnings.warn(
            f"DataFrame尺寸({df_rows}x{df_cols})大于目标区域({required_rows}x{required_cols})，数据将被截断。",
            UserWarning,
        )

    # 以最小代价调整到目标尺寸：优先截断，必要时补 None
    if df_rows != required_rows or df_cols != required_cols:
        df = df.iloc[:required_rows, :required_cols]
        if df.shape != (required_rows, required_cols):
            df = df.reindex(index=range(required_rows), columns=range(required_cols))

    # 预组装写入数据（减少循环内分支）
    data_to_write = df.values.tolist()
    if header and not df.empty:
        col_names = [f"Column_{i}" for i in df.columns] if isinstance(df.columns, pd.RangeIndex) else df.columns.tolist()
        data_to_write = [col_names] + data_to_write
    if index and not df.empty:
        row_names = [f"Row_{i}" for i in df.index] if isinstance(df.index, pd.RangeIndex) else df.index.tolist()
        for i, row in enumerate(data_to_write):
            if header and i == 0:
                data_to_write[i] = [None] + row
            else:
                rn = row_names[i - 1] if header else row_names[i]
                data_to_write[i] = [rn] + row

    # 打开/复用工作簿与工作表
    wb = _workbook or _open_or_create_workbook(excel_name)
    ws = _get_or_create_worksheet(wb, sheet_name)

    # 写入
    cell_set = ws.cell
    sr, sc = start_row, start_col
    for i, row in enumerate(data_to_write):
        r = sr + i
        for j, val in enumerate(row):
            cell_set(row=r, column=sc + j, value=val)

    # 仅在非批量会话时保存
    if _workbook is None and _save:
        wb.save(excel_name)


def _open_or_create_workbook(excel_name: str) -> openpyxl.Workbook:
    try:
        return openpyxl.load_workbook(excel_name, data_only=True, read_only=False)
    except FileNotFoundError:
        wb = openpyxl.Workbook()
        if 'Sheet' in wb.sheetnames:
            wb.remove(wb['Sheet'])
        return wb


def _get_or_create_worksheet(workbook: openpyxl.Workbook, sheet_name: Union[str, int]):
    if isinstance(sheet_name, int):
        if sheet_name < len(workbook.sheetnames):
            return workbook.worksheets[sheet_name]
        return workbook.create_sheet(f"Sheet{sheet_name + 1}")
    return workbook[sheet_name] if sheet_name in workbook.sheetnames else workbook.create_sheet(sheet_name)
