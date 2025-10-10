"""
Excel 读取模块

提供Excel文件范围读取功能
"""

import pandas as pd
import numpy as np
from typing import Any, Optional, Union, List, Dict, Tuple
import warnings


def read_excel_range(file_path: str, 
                     sheet_name: Union[str, int] = 0,
                     ranges: Optional[Union[str, List[str]]] = None,
                     engine: str = 'openpyxl',
                     merge_ranges: bool = False,
                     **kwargs) -> Union[pd.DataFrame, Dict[str, pd.DataFrame]]:
    """
    读取Excel文件的指定范围
    
    参数:
    file_path: Excel文件路径
    sheet_name: 工作表名称或索引，默认为0（第一个工作表）
    ranges: 单个范围字符串或范围列表，如 "A1:C10" 或 ["A1:C10", "E1:G10"]
    engine: 读取引擎，默认 'openpyxl'，也可用 'calamine' 等
    merge_ranges: 是否纵向合并多个范围，默认False返回字典
    **kwargs: 传递给 pd.read_excel 的其他参数
    
    返回:
    - 单个范围：返回 DataFrame
    - 多个范围且 merge_ranges=False：返回字典 {range: DataFrame}
    - 多个范围且 merge_ranges=True：返回合并后的 DataFrame
    
    示例:
        >>> import xlgrab
        >>> 
        >>> # 读取单个范围
        >>> df = xlgrab.read_excel_range("data.xlsx", ranges="A1:C10")
        >>> 
        >>> # 读取多个范围（返回字典）
        >>> dfs = xlgrab.read_excel_range("data.xlsx", ranges=["A1:C10", "E1:G10"])
        >>> 
        >>> # 读取多个范围并合并
        >>> df = xlgrab.read_excel_range("data.xlsx", ranges=["A1:C10", "E1:G10"], merge_ranges=True)
        >>> 
        >>> # 指定工作表
        >>> df = xlgrab.read_excel_range("data.xlsx", sheet_name="Sheet1", ranges="A1:C10")
    """
    
    # 如果没有指定范围，使用标准 read_excel
    if ranges is None:
        return pd.read_excel(file_path, sheet_name=sheet_name, engine=engine, **kwargs)
    
    # 统一转换为列表格式
    range_list = [ranges] if isinstance(ranges, str) else ranges
    
    # 解析范围函数（使用 openpyxl.utils）
    def parse_range(cell_range: str) -> Dict:
        """解析单元格范围，如 A1:C10"""
        from openpyxl.utils import range_boundaries, get_column_letter
        
        try:
            # range_boundaries 返回 (min_col, min_row, max_col, max_row)
            min_col, min_row, max_col, max_row = range_boundaries(cell_range)
            
            # 转换列索引为列字母
            start_col = get_column_letter(min_col)
            end_col = get_column_letter(max_col)
            
            return {
                'start_col': start_col,
                'end_col': end_col,
                'start_row': min_row,
                'end_row': max_row,
                'usecols': f"{start_col}:{end_col}",
                'skiprows': min_row - 1,
                'nrows': max_row - min_row + 1
            }
        except Exception as e:
            raise ValueError(f"无效的范围格式 '{cell_range}': {e}")
    
    # 读取每个范围的数据
    range_data = {}
    
    for cell_range in range_list:
        try:
            range_info = parse_range(cell_range)
            
            # 读取指定范围的数据
            df_range = pd.read_excel(
                file_path,
                sheet_name=sheet_name,
                engine=engine,
                header=None,
                usecols=range_info['usecols'],
                skiprows=range_info['skiprows'],
                nrows=range_info['nrows'],
                **kwargs
            )
            
            range_data[cell_range] = df_range
            
        except Exception as e:
            raise ValueError(f"读取范围 {cell_range} 失败: {e}")
    
    # 返回结果
    if len(range_list) == 1:
        # 单个范围，直接返回 DataFrame
        return range_data[range_list[0]]
    elif merge_ranges:
        # 多个范围，纵向合并
        dfs = list(range_data.values())
        merged_df = pd.concat(dfs, axis=0, ignore_index=True)
        return merged_df
    else:
        # 多个范围，返回字典
        return range_data
