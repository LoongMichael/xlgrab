"""
Excel 合并单元格处理模块

提供解开合并单元格并填充值的功能
"""

import pandas as pd
import numpy as np
from typing import Any, Optional, Union, List, Dict, Tuple
import warnings
import openpyxl


def unmerge_excel(file_path: Union[str, List[str]], 
                  output_path: Optional[Union[str, List[str]]] = None,
                  sheet_names: Optional[Union[str, List[str]]] = None,
                  copy_style: bool = True,
                  verbose: bool = False) -> Dict:
    """
    解开Excel文件中的所有合并单元格并填充值
    
    参数:
    file_path: 输入Excel文件路径或文件路径列表
    output_path: 输出Excel文件路径或路径列表，如果为None则覆盖原文件
    sheet_names: 要处理的工作表名称或名称列表，None表示处理所有工作表
    copy_style: 是否复制单元格格式（数字格式、数据类型等），默认True
    verbose: 是否显示详细处理信息
    
    返回:
    Dict: 处理结果统计
        - total_files: 处理的文件数量
        - total_sheets: 处理的工作表数量
        - total_merged: 处理的合并单元格数量
        - files_info: 各文件的详细信息
    
    示例:
        >>> import xlgrab
        >>> 
        >>> # 处理单个文件的所有工作表
        >>> result = xlgrab.unmerge_excel("input.xlsx", "output.xlsx")
        >>> 
        >>> # 处理单个文件的指定工作表
        >>> result = xlgrab.unmerge_excel("input.xlsx", "output.xlsx", sheet_names="Sheet1")
        >>> result = xlgrab.unmerge_excel("input.xlsx", "output.xlsx", sheet_names=["Sheet1", "Sheet2"])
        >>> 
        >>> # 批量处理多个文件
        >>> files = ["file1.xlsx", "file2.xlsx", "file3.xlsx"]
        >>> result = xlgrab.unmerge_excel(files)
        >>> 
        >>> # 批量处理并指定输出路径
        >>> result = xlgrab.unmerge_excel(files, ["out1.xlsx", "out2.xlsx", "out3.xlsx"])
    """
    
    # 统一转换为列表格式
    file_list = [file_path] if isinstance(file_path, str) else file_path
    
    # 处理输出路径
    if output_path is None:
        output_list = file_list.copy()
    elif isinstance(output_path, str):
        output_list = [output_path]
    else:
        output_list = output_path
    
    # 检查输入输出数量是否匹配
    if len(file_list) != len(output_list):
        raise ValueError(f"输入文件数量 ({len(file_list)}) 与输出路径数量 ({len(output_list)}) 不匹配")
    
    # 处理工作表名称
    if sheet_names is not None:
        sheet_list = [sheet_names] if isinstance(sheet_names, str) else sheet_names
    else:
        sheet_list = None
    
    total_files = 0
    total_sheets = 0
    total_merged = 0
    files_info = []
    
    # 处理每个文件
    for input_file, output_file in zip(file_list, output_list):
        if verbose:
            print(f"\n{'='*60}")
            print(f"处理文件: {input_file}")
            print(f"{'='*60}")
        
        try:
            # 加载工作簿
            workbook = openpyxl.load_workbook(input_file)
            
            # 确定要处理的工作表
            if sheet_list is None:
                sheets_to_process = workbook.sheetnames
            else:
                # 验证工作表是否存在
                sheets_to_process = []
                for sheet_name in sheet_list:
                    if sheet_name in workbook.sheetnames:
                        sheets_to_process.append(sheet_name)
                    elif verbose:
                        print(f"警告: 工作表 '{sheet_name}' 不存在于文件 {input_file}")
            
            file_merged = 0
            sheets_info = []
            
            # 处理每个工作表
            for sheet_name in sheets_to_process:
                worksheet = workbook[sheet_name]
                
                # 调用 unmerge_sheet 处理
                result = unmerge_sheet(worksheet, copy_style=copy_style, verbose=False)
                
                file_merged += result['merged_count']
                sheets_info.append({
                    'sheet_name': sheet_name,
                    'merged_count': result['merged_count']
                })
                
                if verbose and result['merged_count'] > 0:
                    print(f"  工作表 '{sheet_name}': 处理了 {result['merged_count']} 个合并单元格")
            
            # 保存文件
            workbook.save(output_file)
            
            total_files += 1
            total_sheets += len(sheets_info)
            total_merged += file_merged
            
            files_info.append({
                'input_file': input_file,
                'output_file': output_file,
                'sheets_count': len(sheets_info),
                'merged_count': file_merged,
                'sheets_info': sheets_info
            })
            
            if verbose:
                print(f"  已保存到: {output_file}")
                print(f"  共处理 {file_merged} 个合并单元格")
        
        except Exception as e:
            if verbose:
                print(f"  处理失败: {e}")
            files_info.append({
                'input_file': input_file,
                'output_file': output_file,
                'error': str(e)
            })
    
    result = {
        'total_files': total_files,
        'total_sheets': total_sheets,
        'total_merged': total_merged,
        'files_info': files_info
    }
    
    if verbose:
        print(f"\n{'='*60}")
        print(f"全部完成! 共处理 {total_files} 个文件, {total_sheets} 个工作表, {total_merged} 个合并单元格")
        print(f"{'='*60}")
    
    return result


def unmerge_sheet(worksheet, copy_style: bool = True, verbose: bool = False) -> Dict:
    """
    取消单个工作表中的所有合并单元格并填充值
    
    参数:
    worksheet: openpyxl 工作表对象
    copy_style: 是否复制单元格格式（数字格式、数据类型等），默认True
    verbose: 是否显示详细处理信息
    
    返回:
    Dict: 处理结果
        - merged_count: 处理的合并单元格数量
        - merge_details: 每个合并单元格的详细信息
    
    使用示例:
        >>> import openpyxl
        >>> import xlgrab
        >>> wb = openpyxl.load_workbook("input.xlsx")
        >>> ws = wb.active
        >>> result = xlgrab.unmerge_sheet(ws, verbose=True)
        >>> wb.save("output.xlsx")
    """
    
    # 获取所有合并单元格范围
    merged_ranges = list(worksheet.merged_cells.ranges)
    
    if verbose:
        print(f"\n{'='*60}")
        print(f"工作表: {worksheet.title}")
        print(f"发现 {len(merged_ranges)} 个合并单元格")
        print(f"{'='*60}")
    
    if not merged_ranges:
        return {'merged_count': 0, 'merge_details': []}
    
    # 收集所有合并单元格的信息
    merge_info = []
    for i, merged_range in enumerate(merged_ranges, 1):
        # bounds 返回 (min_col, min_row, max_col, max_row)
        min_col, min_row, max_col, max_row = merged_range.bounds
        value = worksheet.cell(min_row, min_col).value
        
        merge_detail = {
            'index': i,
            'range': str(merged_range),
            'bounds': (min_row, min_col, max_row, max_col),
            'value': value,
            'value_type': type(value).__name__
        }
        merge_info.append(merge_detail)
        
        if verbose:
            print(f"\n[{i}] 合并单元格: {merged_range}")
            print(f"    边界: 行{min_row}-{max_row}, 列{min_col}-{max_col}")
            print(f"    原值: '{value}' (类型: {type(value).__name__})")
    
    # 取消所有合并单元格
    if verbose:
        print(f"\n{'='*60}")
        print("开始取消合并...")
        print(f"{'='*60}")
    
    unmerged_count = 0
    for info in merge_info:
        try:
            worksheet.unmerge_cells(info['range'])
            unmerged_count += 1
            if verbose:
                print(f"  ✓ 已取消合并: {info['range']}")
        except ValueError as e:
            if verbose:
                print(f"  ✗ 无法取消合并: {info['range']} - {e}")
            continue
    
    if verbose:
        print(f"\n成功取消 {unmerged_count}/{len(merge_info)} 个合并单元格")
    
    # 填充所有单元格
    if verbose:
        print(f"\n{'='*60}")
        print("开始填充值...")
        print(f"{'='*60}")
    
    fill_count = 0
    for info in merge_info:
        min_row, min_col, max_row, max_col = info['bounds']
        
        # 获取源单元格（合并单元格的左上角）
        source_cell = worksheet.cell(min_row, min_col)
        value = source_cell.value
        
        if verbose:
            print(f"\n[{info['index']}] 填充 {info['range']} 为 '{value}'")
            print(f"    数字格式: {source_cell.number_format}")
            print(f"    数据类型: {source_cell.data_type}")
        
        cells_filled = 0
        for row in range(min_row, max_row + 1):
            for col in range(min_col, max_col + 1):
                target_cell = worksheet.cell(row, col)
                old_value = target_cell.value
                
                # 复制值
                target_cell.value = value
                
                # 复制格式（如果启用）
                if copy_style:
                    if source_cell.number_format:
                        target_cell.number_format = source_cell.number_format
                    if source_cell.data_type:
                        target_cell.data_type = source_cell.data_type
                
                cells_filled += 1
                
                if verbose:
                    from openpyxl.utils import get_column_letter
                    col_letter = get_column_letter(col)
                    if old_value != value:
                        print(f"    {col_letter}{row}: '{old_value}' -> '{value}' (格式: {target_cell.number_format})")
                    else:
                        print(f"    {col_letter}{row}: '{value}' (不变, 格式: {target_cell.number_format})")
        
        fill_count += 1
        if verbose:
            print(f"    共填充 {cells_filled} 个单元格")
    
    if verbose:
        print(f"\n{'='*60}")
        print(f"处理完成! 填充了 {fill_count} 个合并单元格区域")
        print(f"{'='*60}")
    
    return {
        'merged_count': len(merge_info),
        'merge_details': merge_info
    }
