"""
xlgrab工具函数模块
"""

import pandas as pd
import numpy as np
from typing import Any, Optional, Union, List, Dict, Tuple
import warnings
import openpyxl


# ==================== 数据创建工具 ====================
# TODO: 在这里添加数据创建工具函数


# ==================== 数据分析工具 ====================
# TODO: 在这里添加数据分析工具函数


# ==================== 数据转换工具 ====================

def unmerge_excel(file_path: str, output_path: Optional[str] = None) -> None:
    """
    解开Excel文件中的所有合并单元格并填充值
    
    参数:
    file_path: 输入Excel文件路径
    output_path: 输出Excel文件路径，如果为None则覆盖原文件
    
    示例:
        >>> import xlgrab
        >>> xlgrab.unmerge_excel("input.xlsx", "output.xlsx")
        >>> # 或者覆盖原文件
        >>> xlgrab.unmerge_excel("input.xlsx")
    """
    
    if output_path is None:
        output_path = file_path
    
    # 加载工作簿
    workbook = openpyxl.load_workbook(file_path)
    
    # 遍历所有工作表
    for sheet_name in workbook.sheetnames:
        worksheet = workbook[sheet_name]
        
        # 获取所有合并单元格范围
        merged_ranges = list(worksheet.merged_cells.ranges)
        
        # 先收集所有合并单元格的信息
        merge_info = []
        for merged_range in merged_ranges:
            min_row, min_col, max_row, max_col = merged_range.bounds
            value = worksheet.cell(min_row, min_col).value
            merge_info.append({
                'range': str(merged_range),
                'bounds': (min_row, min_col, max_row, max_col),
                'value': value
            })
        
        # 取消所有合并单元格
        for info in merge_info:
            try:
                worksheet.unmerge_cells(info['range'])
            except ValueError:
                # 如果合并单元格已经不存在，跳过
                continue
        
        # 填充所有单元格
        for info in merge_info:
            min_row, min_col, max_row, max_col = info['bounds']
            value = info['value']
            
            if value is not None:
                for row in range(min_row, max_row + 1):
                    for col in range(min_col, max_col + 1):
                        # 使用 cell() 方法获取或创建单元格
                        cell = worksheet.cell(row, col)
                        cell.value = value
    
    # 保存文件
    workbook.save(output_path)
    print(f"处理完成: {output_path}")


# ==================== 数据导出工具 ====================
# TODO: 在这里添加数据导出工具函数